from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
import numpy
import matplotlib.pyplot as plt
import numpy as np
import pylab as p

wb = load_workbook(filename = r'empty_book.xlsx')
ws = wb.get_active_sheet()

dest_filename = input('What would you like to name the output file? ') + '.xlsx'
sheet_ranges = wb.get_sheet_by_name(name = 'range names')

used_rows = input('Which rows of the microtitre plate are your samples in? (write seperated by spaces)')
used_rows_list = used_rows.split()
used_rows_start = used_rows_list[0]
used_rows_end = used_rows_list[-1]
used_rows_from = used_rows_start + '1'
used_rows_to = used_rows_end + '3'
usedrowsrange = used_rows_from + ':' + used_rows_to

mock_reading_raw = input('Which cell of the excel file is your original mock in? ')
mock_reading = (sheet_ranges.cell(mock_reading_raw).value)

end_user_col = input('Which column of the microtitre plate is your last sample in? (Just a single number please) ')
end_user_col_int = int(end_user_col)
end_user_col_taper = -(-end_user_col_int)

assay_time_input = input('How many minutes was your assay for? ')
assay_time = int(assay_time_input)

dilution_cells_input = input('By what factor did you dilute the cells? ')
dilution_cells = int(dilution_cells_input)

sheet_ranges = wb.get_sheet_by_name(name = 'range names')

originlist = []
for row in ws.range(usedrowsrange):
    for cell in row:
        originlist.append(cell.value)
originlist = list(range(1, 25))
end_user_col_taper = 12-end_user_col_int
originlistslice = originlist[0:end_user_col_taper]

originlist_array = numpy.array(originlistslice)

minus_mock = originlist_array-mock_reading
print(minus_mock)
by_time = minus_mock/assay_time
print(by_time)
by_dilution = by_time*dilution_cells
print(by_time)
by_30ul = by_dilution/30
print(by_30ul1)
into_ml = by_30ul1*1000
print(into_ml)


#fig = p.figure()
#ax = fig.add_subplot(1,1,1)
#x = [1,2,3,4,5,6]
#y = originlist_array
#ax.bar(x,y,facecolor='#777777', align='center')
#ax.set_ylabel('Secretion')
#ax.set_title('Sample Secretions',fontstyle='italic')
#ax.set_xticks(x)
#group_labels = range(1-3)
#ax.set_xticklabels(group_labels)
#fig.autofmt_xdate()
#ax.set_ylim([0,50])
#p.show()



#ws = wb.create_sheet()
#ws.title = "PyOutput" #Now the new sheet is editable
#ws.range('A1:C2').value = startval
#wb.save(filename = dest_filename)
